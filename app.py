# app.py → HR Insight Bot (FINAL VERSION - PDF WORKS 100%)
# Run: python app.py

import joblib
import pandas as pd
import numpy as np
import sqlite3
import datetime
import io
import bcrypt
import openpyxl
import matplotlib.pyplot as plt
from fpdf import FPDF
from flask import Flask, request, render_template_string, session, redirect, url_for, send_file, flash
from textblob import TextBlob
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
import traceback

app = Flask(__name__)
app.secret_key = 'hr_insight_bot_2025_secret'

# -------------------------- Database --------------------------
DB = 'employee_db.db'
conn = sqlite3.connect(DB, check_same_thread=False)
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS employee (
             id TEXT PRIMARY KEY, name TEXT, age INT, income INT, sat INT,
             overtime TEXT, involve INT, feedback TEXT, leaves_taken INT, password_hash TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS task (
             emp_id TEXT, task TEXT, status TEXT, ts TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS chat (
             emp_id TEXT, role TEXT, message TEXT, ts TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS applications (
             id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, designation TEXT,
             experience TEXT, role TEXT, ts TEXT)''')
try:
    c.execute("ALTER TABLE employee ADD COLUMN password_hash TEXT")
except:
    pass
conn.commit()

# -------------------------- Excel Export --------------------------
EXCEL_FILE = 'hr_data.xlsx'

def init_excel():
    try:
        pd.read_excel(EXCEL_FILE)
    except FileNotFoundError:
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            pd.DataFrame(columns=['emp_id','name','age','income','sat','overtime','involve','feedback','leaves_taken','ts']).to_excel(writer, sheet_name='Employees', index=False)
            pd.DataFrame(columns=['name','designation','experience','role','ts']).to_excel(writer, sheet_name='Applicants', index=False)

def append_employee_row(row):
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name='Employees')
    except:
        df = pd.DataFrame(columns=['emp_id','name','age','income','sat','overtime','involve','feedback','leaves_taken','ts'])
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Employees', index=False)

def append_applicant_row(row):
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name='Applicants')
    except:
        df = pd.DataFrame(columns=['name','designation','experience','role','ts'])
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Applicants', index=False)

init_excel()

# -------------------------- Job Roles --------------------------
JOB_ROLES = [
    "Data Scientist", "Software Engineer", "HR Manager", "Product Manager",
    "DevOps Engineer", "UX Designer", "Business Analyst", "QA Engineer"
]

# ------------------- 14 FEATURES -------------------
FEATURES_14 = [
    'Age', 'MonthlyIncome', 'JobSatisfaction', 'JobInvolvement',
    'YearsAtCompany', 'YearsInCurrentRole', 'YearsWithCurrManager',
    'TotalWorkingYears', 'DistanceFromHome', 'WorkLifeBalance',
    'EnvironmentSatisfaction', 'FeedbackSentiment',
    'OverTime_Yes', 'OverTime_No'
]

# -------------------------- DEFAULTS --------------------------
DEFAULTS = {
    'Age': 37, 'MonthlyIncome': 6500, 'JobSatisfaction': 3, 'JobInvolvement': 3,
    'YearsAtCompany': 6, 'YearsInCurrentRole': 4, 'YearsWithCurrManager': 4,
    'TotalWorkingYears': 11, 'DistanceFromHome': 9, 'WorkLifeBalance': 3,
    'EnvironmentSatisfaction': 3, 'OverTime': 'No'
}

# -------------------------- Self-Train Models --------------------------
def train_models():
    np.random.seed(42)
    n_samples = 1000
    data = {
        'Age': np.random.randint(18, 65, n_samples),
        'MonthlyIncome': np.random.randint(3000, 20000, n_samples),
        'JobSatisfaction': np.random.randint(1, 5, n_samples),
        'JobInvolvement': np.random.randint(1, 5, n_samples),
        'YearsAtCompany': np.random.randint(0, 40, n_samples),
        'YearsInCurrentRole': np.random.randint(0, 18, n_samples),
        'YearsWithCurrManager': np.random.randint(0, 17, n_samples),
        'TotalWorkingYears': np.random.randint(0, 40, n_samples),
        'DistanceFromHome': np.random.randint(1, 30, n_samples),
        'WorkLifeBalance': np.random.randint(1, 5, n_samples),
        'EnvironmentSatisfaction': np.random.randint(1, 5, n_samples),
        'OverTime': np.random.choice(['Yes', 'No'], n_samples),
        'FeedbackSentiment': np.random.uniform(-1, 1, n_samples)
    }
    df = pd.DataFrame(data)
    df['OverTime_Yes'] = (df['OverTime'] == 'Yes').astype(int)
    df['OverTime_No'] = (df['OverTime'] == 'No').astype(int)
    df['Attrition'] = ((df['JobSatisfaction'] <= 2) & (df['MonthlyIncome'] < 6000)).astype(int)
    df['Promotion'] = ((df['JobSatisfaction'] >= 3) & (df['TotalWorkingYears'] > 5)).astype(int)

    X = df[FEATURES_14]
    y_attrition = df['Attrition']
    y_promotion = df['Promotion']

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    pca = PCA(n_components=10)
    X_pca = pca.fit_transform(X_scaled)

    model_attrition = RandomForestClassifier(n_estimators=100, random_state=42)
    model_promotion = RandomForestClassifier(n_estimators=100, random_state=42)
    model_attrition.fit(X_pca, y_attrition)
    model_promotion.fit(X_pca, y_promotion)

    return model_attrition, model_promotion, scaler, pca

model_attrition, model_promotion, scaler, pca = train_models()

# -------------------------- Password Utils --------------------------
def hash_password(pw):
    return bcrypt.hashpw(pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def check_password(pw, hashed):
    return bcrypt.checkpw(pw.encode('utf-8'), hashed.encode('utf-8'))

# -------------------------- Prediction --------------------------
def predict(features):
    full = DEFAULTS.copy()
    full.update(features)
    df = pd.DataFrame([full])
    df['FeedbackSentiment'] = TextBlob(full.get('Feedback', '')).sentiment.polarity
    df['OverTime_Yes'] = (df['OverTime'] == 'Yes').astype(int)
    df['OverTime_No'] = (df['OverTime'] == 'No').astype(int)
    X = df[FEATURES_14]
    X_scaled = scaler.transform(X)
    X_pca = pca.transform(X_scaled)

    attrition_pred = model_attrition.predict(X_pca)[0]
    attrition_prob = model_attrition.predict_proba(X_pca)[0][1]
    promotion_pred = model_promotion.predict(X_pca)[0]
    promotion_prob = model_promotion.predict_proba(X_pca)[0][1]

    return (
        ('Yes' if attrition_pred == 1 else 'No'), round(float(attrition_prob), 3),
        ('Yes' if promotion_pred == 1 else 'No'), round(float(promotion_prob), 3)
    )

# -------------------------- CHART GENERATORS --------------------------
def create_gauge_chart(value, title):
    fig, ax = plt.subplots(figsize=(3, 2), subplot_kw=dict(polar=True))
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)
    ax.set_ylim(0, 1)
    ax.set_yticks([])
    ax.grid(False)
    ax.spines['polar'].set_visible(False)

    color = '#dc3545' if value > 0.7 else '#ffaa00' if value > 0.4 else '#28a745'
    angle = value * np.pi
    ax.barh(1, angle, color=color, height=0.3)
    ax.text(0, 0.3, f"{value:.0%}", ha='center', va='center', fontsize=12, fontweight='bold', color='white')
    ax.set_title(title, pad=15, fontsize=9)
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=100, transparent=True)
    plt.close()
    buf.seek(0)
    return buf

def create_task_pie(tasks):
    done = sum(1 for _, s in tasks if s == 'Done')
    pending = len(tasks) - done
    if done + pending == 0:
        return None
    labels = ['Done', 'Pending']
    sizes = [done, pending]
    colors = ['#28a745', '#dc3545']

    fig, ax = plt.subplots(figsize=(3, 3))
    ax.pie(sizes, labels=labels, colors=colors, autopct='%1.0f%%', startangle=90)
    ax.axis('equal')
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=100)
    plt.close()
    buf.seek(0)
    return buf

# -------------------------- PDF Report (100% SAFE) --------------------------
# -------------------------- PDF Report (FINAL FIX) --------------------------
class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, 'HR Insight Report', ln=True, align='C')
        self.ln(5)

    def add_image_stream(self, stream):
        if stream:
            stream.seek(0)
            try:
                self.image(stream, w=60, h=45)
            except:
                pass  # Skip image if broken

def generate_pdf(emp_id, profile_dict, results, tasks):
    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Employee Info
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Employee Report", ln=True)
    pdf.ln(5)

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Employee Information", ln=True)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 7, f"ID: {emp_id}", ln=True)
    pdf.cell(0, 7, f"Name: {profile_dict.get('name', 'N/A')}", ln=True)
    pdf.cell(0, 7, f"Age: {profile_dict.get('age', 'N/A')}", ln=True)
    pdf.cell(0, 7, f"Monthly Income: INR {profile_dict.get('income', 'N/A')}", ln=True)
    pdf.cell(0, 7, f"Job Satisfaction: {profile_dict.get('sat', 'N/A')}/4", ln=True)
    pdf.cell(0, 7, f"Overtime: {profile_dict.get('overtime', 'N/A')}", ln=True)
    pdf.cell(0, 7, f"Job Involvement: {profile_dict.get('involve', 'N/A')}/4", ln=True)
    feedback = (profile_dict.get('feedback') or '')[:100]
    pdf.cell(0, 7, f"Feedback: {feedback}{'...' if len(feedback) >= 100 else ''}", ln=True)
    pdf.ln(5)

    # Predictions
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Predictions", ln=True)
    attrition_prob = results.get('attrition_prob', 0)
    promotion_prob = results.get('promotion_prob', 0)

    risk = "High" if attrition_prob > 0.7 else "Medium" if attrition_prob > 0.4 else "Low"
    pdf.cell(90, 8, f"Attrition: {results.get('attrition','N/A')} ({attrition_prob:.1%}) - {risk}", ln=0)
    chart1 = create_gauge_chart(attrition_prob, "Risk")
    pdf.add_image_stream(chart1)
    pdf.ln(50)

    pdf.cell(90, 8, f"Promotion: {results.get('promotion','N/A')} ({promotion_prob:.1%})", ln=0)
    chart2 = create_gauge_chart(promotion_prob, "Chance")
    pdf.add_image_stream(chart2)
    pdf.ln(55)

    # Tasks
    if tasks:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Task Tracker", ln=True)
        chart3 = create_task_pie(tasks)
        if chart3:
            pdf.add_image_stream(chart3)
        pdf.ln(75)
        pdf.set_font("Arial", "", 10)
        for task, status in tasks[:10]:
            pdf.cell(0, 6, f"- {task} [{status}]", ln=True)
        if len(tasks) > 10:
            pdf.cell(0, 6, f"... and {len(tasks)-10} more", ln=True)
    else:
        pdf.cell(0, 8, "No tasks recorded.", ln=True)

    buffer = io.BytesIO()
    try:
        # Use file-based output to avoid PNG crash
        temp_path = f"temp_report_{emp_id}.pdf"
        pdf.output(temp_path)
        with open(temp_path, 'rb') as f:
            buffer.write(f.read())
        import os
        os.remove(temp_path)  # Clean up
        buffer.seek(0)
        return buffer
    except Exception as e:
        print("PDF SAVE ERROR:", e)
        # Fallback: Text-only PDF
        pdf = PDF()
        pdf.add_page()
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 10, "PDF generation failed. Showing text only.", ln=True)
        pdf.cell(0, 10, f"Employee: {profile_dict.get('name')}", ln=True)
        pdf.cell(0, 10, f"Attrition: {results.get('attrition')} ({attrition_prob:.1%})", ln=True)
        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)
        return buffer

# -------------------------- Chatbot --------------------------
def get_employee_bot_response(name, leaves_taken, user_msg):
    msg = user_msg.lower().strip()
    if any(g in msg for g in ["hi","hello","hey"]):
        return f"Hello {name}. How may I assist you today?"
    if "leave" in msg:
        return f"You have used {leaves_taken} leaves. Remaining: {30-leaves_taken} out of 30."
    if "salary" in msg or "pay" in msg:
        return "Salary is credited on the 1st of every month. Check HR portal for payslip."
    if "promotion" in msg:
        return "Promotions are based on performance and tenure."
    if "task" in msg:
        return "Manage tasks in the Task Tracker section."
    if "report" in msg:
        return "Click 'Download Full PDF Report' after predictions."
    return "Ask about leaves, salary, tasks, or reports."

# -------------------------- HTML TEMPLATES --------------------------
HTML_WELCOME = """
<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><title>HR Insight Bot</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<style>body{background:linear-gradient(135deg,#667eea,#764ba2);color:white;height:100vh;display:flex;align-items:center;justify-content:center;}
.btn-lg{padding:1rem 2rem;font-size:1.2rem;border-radius:12px;}</style>
</head><body>
<div class="container text-center">
<h1 class="display-4">Welcome to HR Dashboard</h1>
<p class="lead">November 16, 2025 | India</p>
<div class="row mt-5 justify-content-center">
<div class="col-5"><a href="/employee_login_page" class="btn btn-light btn-lg w-100">Employee Login</a></div>
<div class="col-5"><a href="/applicant" class="btn btn-outline-light btn-lg w-100">New Applicant</a></div>
</div></div></body></html>
"""

HTML_LOGIN = """
<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><title>Employee Login</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<style>body{background:linear-gradient(135deg,#667eea,#764ba2);color:white;height:100vh;display:flex;align-items:center;justify-content:center;}
.card{background:rgba(255,255,255,0.15);backdrop-filter:blur(10px);border:none;border-radius:15px;}</style>
</head><body>
<div class="container">
<div class="card p-4 col-md-6 mx-auto">
<h3 class="text-center">Employee Login</h3>
<form method="post" action="/employee_login">
<div class="mb-3"><input name="emp_id" class="form-control form-control-lg" placeholder="Employee ID" required/></div>
<div class="mb-3"><input name="password" type="password" class="form-control form-control-lg" placeholder="Password" required/></div>
<button class="btn btn-primary w-100">Login</button>
</form>
<div class="text-center mt-3">
  <a href="/forgot_password" class="text-warning">Forgot Password?</a>
</div>
<a href="/" class="btn btn-link text-light mt-2 d-block text-center">Back</a>
</div></div></body></html>
"""

HTML_FORGOT_PASSWORD = """
<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><title>Forgot Password</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<style>body{background:linear-gradient(135deg,#667eea,#764ba2);color:white;height:100vh;display:flex;align-items:center;justify-content:center;}
.card{background:rgba(255,255,255,0.15);backdrop-filter:blur(10px);border:none;border-radius:15px;}</style>
</head><body>
<div class="container">
<div class="card p-4 col-md-6 mx-auto">
<h3 class="text-center">Password Recovery</h3>
<p class="text-light text-center">Enter your Employee ID to reset password</p>
<form method="post" action="/recover_password">
<div class="mb-3"><input name="emp_id" class="form-control form-control-lg" placeholder="Employee ID" required/></div>
<button class="btn btn-warning w-100">Continue</button>
</form>
<a href="/employee_login_page" class="btn btn-link text-light mt-2 d-block text-center">Back to Login</a>
</div></div></body></html>
"""

HTML_SET_PASSWORD = """
<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><title>Set New Password</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<style>body{background:linear-gradient(135deg,#667eea,#764ba2);color:white;height:100vh;display:flex;align-items:center;justify-content:center;}
.card{background:rgba(255,255,255,0.15);backdrop-filter:blur(10px);border:none;border-radius:15px;}</style>
</head><body>
<div class="container">
<div class="card p-4 col-md-6 mx-auto">
<h3 class="text-center">Set New Password</h3>
<p class="text-light text-center">For <strong>{{ emp_id }}</strong></p>
<form method="post" action="/set_password">
<input type="hidden" name="emp_id" value="{{ emp_id }}"/>
<div class="mb-3"><input name="password" type="password" class="form-control form-control-lg" placeholder="New Password" required/></div>
<div class="mb-3"><input name="confirm" type="password" class="form-control form-control-lg" placeholder="Confirm Password" required/></div>
<button class="btn btn-success w-100">Set & Login</button>
</form>
<a href="/employee_login_page" class="btn btn-link text-light mt-2 d-block text-center">Back</a>
</div></div></body></html>
"""

HTML_EMPLOYEE = """
<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><title>HR Insight Bot</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
<style>body{background:linear-gradient(135deg,#667eea,#764ba2);color:white;}
.card{background:rgba(255,255,255,0.15);backdrop-filter:blur(10px);border:none;border-radius:15px;}
.result{padding:12px;border-radius:10px;font-weight:bold;margin:8px 0;text-align:center;}
.high{background:#dc3545;color:#fff;}.medium{background:#ffaa00;color:#fff;}.low{background:#28a745;color:#fff;}
.chat-bubble{padding:10px;border-radius:15px;margin:5px 0;max-width:80%;}
.user{background:#fff;color:#333;align-self:flex-end;}.bot{background:#e0e0e0;color:#333;}
</style>
</head><body class="p-3 p-md-5">
<div class="container">
<h1 class="text-center mb-2 display-5">HR Insight Bot</h1>
<p class="text-center text-light">November 16, 2025 | India</p>
<div class="text-end mb-3"><strong>{{ session.emp_id }}</strong> | <a href="/" class="text-warning">Logout</a></div>

{% with messages = get_flashed_messages() %}
  {% if messages %}<div class="alert alert-{{ 'danger' if 'failed' in messages[0].lower() else 'success' }}">{{ messages[0] }}</div>{% endif %}
{% endwith %}

<div class="row g-4">
  <div class="col-lg-4">
    <div class="card p-3">
      <h4>Update Profile</h4>
      <form method="post" action="/save_profile">
        <input type="hidden" name="emp_id" value="{{ session.emp_id }}"/>
        <div class="mb-2"><label>Name</label><input name="name" class="form-control" value="{{ profile.name or '' }}"/></div>
        <div class="mb-2"><label>Age</label><input name="age" type="number" class="form-control" value="{{ profile.age or 30 }}"/></div>
        <div class="mb-2"><label>Monthly Income (INR)</label><input name="income" type="number" class="form-control" value="{{ profile.income or 50000 }}"/></div>
        <div class="mb-2"><label>Job Satisfaction (1-4)</label><input name="sat" type="number" min="1" max="4" class="form-control" value="{{ profile.sat or 3 }}"/></div>
        <div class="mb-2"><label>Overtime</label>
          <select name="overtime" class="form-select">
            <option value="No" {% if (profile.overtime or 'No')=='No' %}selected{% endif %}>No</option>
            <option value="Yes" {% if profile.overtime=='Yes' %}selected{% endif %}>Yes</option>
          </select>
        </div>
        <div class="mb-2"><label>Job Involvement (1-4)</label><input name="involve" type="number" min="1" max="4" class="form-control" value="{{ profile.involve or 3 }}"/></div>
        <div class="mb-3"><label>Feedback</label><textarea name="feedback" class="form-control" rows="2">{{ profile.feedback or '' }}</textarea></div>
        <button class="btn btn-warning w-100">Save & Predict</button>
      </form>
    </div>
  </div>

  <div class="col-lg-4">
    <div class="card p-3">
      <h4>Predictions</h4>
      {% if results %}
        <div class="result {{ 'high' if results.attrition_prob>0.7 else 'medium' if results.attrition_prob>0.4 else 'low' }}">
          Attrition: <strong>{{ results.attrition }}</strong> ({{ "%.1f"|format(results.attrition_prob*100) }}%)
        </div>
        <div class="result {{ 'high' if results.promotion=='Yes' else 'low' }}">
          Promotion: <strong>{{ results.promotion }}</strong> ({{ "%.1f"|format(results.promotion_prob*100) }}%)
        </div>
        <form method="post" action="/download_pdf" class="mt-3">
          <input type="hidden" name="emp_id" value="{{ session.emp_id }}"/>
          <button class="btn btn-success w-100">Download Full PDF Report</button>
        </form>
      {% else %}
        <p class="text-muted">Click “Save & Predict” to generate results.</p>
      {% endif %}
    </div>
  </div>

  <div class="col-lg-4">
    <div class="card p-3 mb-3">
      <h4>Task Tracker</h4>
      <form method="post" action="/add_task" class="input-group mb-2">
        <input type="hidden" name="emp_id" value="{{ session.emp_id }}"/>
        <input name="task" class="form-control" placeholder="New task"/>
        <button class="btn btn-success">Add</button>
      </form>
      <div id="taskChart" style="height:150px;"></div>
      <ul class="list-group mt-2">
        {% for t in tasks %}
          <li class="list-group-item d-flex justify-content-between">
            {{ t.task }}
            {% if t.status == 'Pending' %}
              <form method="post" action="/complete_task" class="d-inline">
                <input type="hidden" name="emp_id" value="{{ session.emp_id }}"/>
                <input type="hidden" name="task" value="{{ t.task }}"/>
                <button class="btn btn-success btn-sm">Done</button>
              </form>
            {% else %}
              <span class="badge bg-success">Done</span>
            {% endif %}
          </li>
        {% endfor %}
      </ul>
    </div>

    <div class="card p-3">
      <h4>Chat with HR Bot</h4>
      <div style="height:200px;overflow-y:auto;display:flex;flex-direction:column;" id="chatBox">
        {% for msg in chat_history %}
          <div class="chat-bubble {{ 'user' if msg.role=='user' else 'bot' }} align-self-{{ 'end' if msg.role=='user' else 'start' }}">
            {{ msg.message }}
          </div>
        {% endfor %}
      </div>
      <form method="post" action="/chat" class="mt-2 d-flex">
        <input type="hidden" name="emp_id" value="{{ session.emp_id }}"/>
        <input name="message" class="form-control form-control-sm me-1" placeholder="Ask anything..."/>
        <button class="btn btn-primary btn-sm">Send</button>
      </form>
    </div>
  </div>
</div>

<script>
  {% if tasks %}
    const done = {{ tasks | selectattr("status","equalto","Done") | list | length }};
    const pending = {{ tasks | selectattr("status","equalto","Pending") | list | length }};
    if (done + pending > 0) {
      Plotly.newPlot('taskChart', [{
        values: [done, pending],
        labels: ['Done', 'Pending'],
        marker: {colors: ['#28a745', '#dc3545']},
        type: 'pie', hole: 0.5
      }], {height: 150, margin: {t:0,b:0,l:0,r:0}});
    }
  {% endif %}
</script>
</body></html>
"""

HTML_APPLICANT = """
<!DOCTYPE html>
<html><head><meta charset="UTF-8"/><title>HR Insight Bot - Applicant</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet"/>
<style>
body{background:linear-gradient(135deg,#667eea,#764ba2);color:white;}
.card{background:rgba(255,255,255,0.15);backdrop-filter:blur(10px);border:none;border-radius:15px;}
</style>
</head><body class="p-5">
<div class="container">
<h1 class="text-center mb-4">New Applicant Portal</h1>
<a href="/" class="btn btn-outline-light mb-3">Back</a>
<div class="row g-4">
  <div class="col-lg-6">
    <div class="card p-4">
      <h4>Apply for Job</h4>
      {% with messages = get_flashed_messages() %}
        {% if messages %}<div class="alert alert-success">{{ messages[0] }}</div>{% endif %}
      {% endwith %}
      <form method="post" action="/submit_application">
        <div class="mb-3"><label>Full Name</label><input name="name" class="form-control" required/></div>
        <div class="mb-3"><label>Current Designation</label><input name="designation" class="form-control" required/></div>
        <div class="mb-3"><label>Years of Experience</label><input name="experience" type="number" min="0" class="form-control" required/></div>
        <div class="mb-3"><label>Apply For Role</label>
          <select name="role" class="form-select" required>
            <option value="">-- Select Role --</option>
            {% for role in job_roles %}
              <option value="{{ role }}">{{ role }}</option>
            {% endfor %}
          </select>
        </div>
        <button class="btn btn-success w-100">Submit Application</button>
      </form>
    </div>
  </div>
  <div class="col-lg-6">
    <div class="card p-4">
      <h4>Chat with HR Bot</h4>
      <div style="height:400px;overflow-y:auto;display:flex;flex-direction:column;" id="chatBox">
        {% if not chat_history %}
          <div class="chat-bubble bot">Hello! Welcome to HR Service Chatbot.</div>
        {% endif %}
        {% for msg in chat_history %}
          <div class="chat-bubble {{ 'user' if msg.role == 'user' else 'bot' }} align-self-{{ 'end' if msg.role == 'user' else 'start' }}">
            {{ msg.message }}
          </div>
        {% endfor %}
      </div>
      <div class="options-bar mt-2 text-center">
        <button class="btn btn-light btn-sm me-1" onclick="send('Job roles')">Job Roles</button>
        <button class="btn btn-light btn-sm me-1" onclick="send('Vacancies')">Vacancies</button>
        <button class="btn btn-light btn-sm" onclick="send('Guidelines')">Guidelines</button>
      </div>
      <form method="post" action="/applicant_chat" id="chatForm" class="d-none">
        <input name="message" id="msgInput"/>
      </form>
    </div>
  </div>
</div>
</div>
<script>
function send(opt) {
  document.getElementById('msgInput').value = opt;
  document.getElementById('chatForm').submit();
}
</script>
</body></html>
"""

# -------------------------- Routes --------------------------
@app.route('/')
def welcome():
    return render_template_string(HTML_WELCOME)

@app.route('/employee_login_page')
def employee_login_page():
    return render_template_string(HTML_LOGIN)

@app.route('/forgot_password')
def forgot_password():
    return render_template_string(HTML_FORGOT_PASSWORD)

@app.route('/recover_password', methods=['POST'])
def recover_password():
    emp_id = request.form.get('emp_id', '').strip()
    if not emp_id:
        flash("Please enter your Employee ID.")
        return redirect(url_for('forgot_password'))
    session['pending_emp_id'] = emp_id
    return render_template_string(HTML_SET_PASSWORD, emp_id=emp_id)

@app.route('/employee_login', methods=['POST'])
def employee_login():
    try:
        emp_id = request.form.get('emp_id', '').strip()
        password = request.form.get('password', '').strip()
        if not emp_id or not password:
            flash("Both fields are required.")
            return redirect(url_for('employee_login_page'))

        c.execute("SELECT password_hash FROM employee WHERE id=?", (emp_id,))
        row = c.fetchone()

        if row is None or row[0] is None:
            session['pending_emp_id'] = emp_id
            return render_template_string(HTML_SET_PASSWORD, emp_id=emp_id)

        if check_password(password, row[0]):
            session.clear()
            session['emp_id'] = emp_id
            return redirect(url_for('employee_dashboard'))
        else:
            flash("Invalid password.")
            return redirect(url_for('employee_login_page'))
    except Exception:
        flash("Login failed.")
        return redirect(url_for('employee_login_page'))

@app.route('/set_password', methods=['POST'])
def set_password():
    try:
        emp_id = request.form['emp_id']
        pw1 = request.form['password']
        pw2 = request.form['confirm']
        if pw1 != pw2:
            flash("Passwords do not match.")
            return render_template_string(HTML_SET_PASSWORD, emp_id=emp_id)
        if len(pw1) < 4:
            flash("Password must be at least 4 characters.")
            return render_template_string(HTML_SET_PASSWORD, emp_id=emp_id)

        hashed = hash_password(pw1)
        c.execute('''INSERT OR REPLACE INTO employee 
                     (id, name, age, income, sat, overtime, involve, feedback, leaves_taken, password_hash)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (emp_id, f"Employee {emp_id}", 30, 50000, 3, "No", 3, "", 0, hashed))
        conn.commit()
        session.clear()
        session['emp_id'] = emp_id
        flash("Password set successfully! Welcome!")
        return redirect(url_for('employee_dashboard'))
    except Exception:
        flash("Error saving password.")
        return redirect(url_for('employee_login_page'))

@app.route('/employee/dashboard')
def employee_dashboard():
    if 'emp_id' not in session:
        return redirect(url_for('employee_login_page'))

    emp_id = session['emp_id']
    c.execute("SELECT * FROM employee WHERE id=?", (emp_id,))
    profile = c.fetchone()
    if not profile:
        c.execute('''INSERT INTO employee (id, name, age, income, sat, overtime, involve, feedback, leaves_taken)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (emp_id, f"Employee {emp_id}", 30, 50000, 3, "No", 3, "", 0))
        conn.commit()
        c.execute("SELECT * FROM employee WHERE id=?", (emp_id,))
        profile = c.fetchone()

    profile_dict = dict(zip(['id','name','age','income','sat','overtime','involve','feedback','leaves_taken','password_hash'], profile))

    c.execute("SELECT task, status FROM task WHERE emp_id=?", (emp_id,))
    tasks = c.fetchall()
    tasks = [type('obj', (object,), {'task': r[0], 'status': r[1]}) for r in tasks]

    c.execute("SELECT role, message FROM chat WHERE emp_id=? ORDER BY ts", (emp_id,))
    chat_history = [type('obj', (object,), {'role': r[0], 'message': r[1]}) for r in c.fetchall()]

    results = session.get('results', {})
    return render_template_string(HTML_EMPLOYEE, profile=profile_dict, tasks=tasks, results=results, chat_history=chat_history, job_roles=JOB_ROLES)

@app.route('/applicant')
def applicant():
    session['emp_id'] = 'applicant_' + str(datetime.datetime.now().timestamp())
    return redirect(url_for('applicant_portal'))

@app.route('/applicant/portal')
def applicant_portal():
    emp_id = session['emp_id']
    c.execute("SELECT role, message FROM chat WHERE emp_id=? ORDER BY ts", (emp_id,))
    chat_history = [type('obj', (object,), {'role': r[0], 'message': r[1]}) for r in c.fetchall()]
    return render_template_string(HTML_APPLICANT, chat_history=chat_history, job_roles=JOB_ROLES)

@app.route('/applicant_chat', methods=['POST'])
def applicant_chat_post():
    emp_id = session['emp_id']
    user_msg = request.form['message']
    c.execute("INSERT INTO chat VALUES (?,?,?,?)", (emp_id, 'user', user_msg, datetime.datetime.now().isoformat()))
    reply = "Available Roles:\n• " + "\n• ".join(JOB_ROLES) if "job" in user_msg.lower() else \
            "Openings: 3+ roles." if "vacanc" in user_msg.lower() else \
            "Guidelines:\n• 30 days leave\n• Hybrid work" if "guide" in user_msg.lower() else \
            "Choose: Job roles, Vacancies, Guidelines."
    c.execute("INSERT INTO chat VALUES (?,?,?,?)", (emp_id, 'bot', reply, datetime.datetime.now().isoformat()))
    conn.commit()
    return redirect(url_for('applicant_portal'))

@app.route('/submit_application', methods=['POST'])
def submit_application():
    try:
        name = request.form['name'].strip()
        designation = request.form['designation'].strip()
        experience = request.form['experience']
        role = request.form['role']
        if not all([name, designation, experience, role]):
            flash("All fields are required.")
            return redirect(url_for('applicant_portal'))
        c.execute("INSERT INTO applications (name, designation, experience, role, ts) VALUES (?,?,?,?,?)",
                  (name, designation, experience, role, datetime.datetime.now().isoformat()))
        conn.commit()
        append_applicant_row({'name': name, 'designation': designation, 'experience': experience, 'role': role, 'ts': datetime.datetime.now().isoformat()})
        flash(f"Application for {role} submitted.")
    except Exception:
        flash("Error submitting application.")
    return redirect(url_for('applicant_portal'))

@app.route('/save_profile', methods=['POST'])
def save_profile():
    try:
        data = request.form
        emp_id = data['emp_id']
        c.execute('''INSERT OR REPLACE INTO employee 
                     (id, name, age, income, sat, overtime, involve, feedback, leaves_taken, password_hash)
                     SELECT id, ?, ?, ?, ?, ?, ?, ?, leaves_taken, password_hash FROM employee WHERE id=?''',
                  (data['name'], int(data['age']), int(data['income']), int(data['sat']),
                   data['overtime'], int(data['involve']), data['feedback'], emp_id))
        conn.commit()

        leaves_taken = c.execute("SELECT leaves_taken FROM employee WHERE id=?", (emp_id,)).fetchone()[0]
        append_employee_row({
            'emp_id': emp_id, 'name': data['name'], 'age': int(data['age']), 'income': int(data['income']),
            'sat': int(data['sat']), 'overtime': data['overtime'], 'involve': int(data['involve']),
            'feedback': data['feedback'], 'leaves_taken': leaves_taken, 'ts': datetime.datetime.now().isoformat()
        })

        attrition, attrition_prob, promotion, promotion_prob = predict({
            'Age': int(data['age']), 'MonthlyIncome': int(data['income']),
            'JobSatisfaction': int(data['sat']), 'OverTime': data['overtime'],
            'JobInvolvement': int(data['involve']), 'Feedback': data['feedback']
        })

        session['results'] = {
            'attrition': attrition, 'attrition_prob': attrition_prob,
            'promotion': promotion, 'promotion_prob': promotion_prob
        }
        flash("Profile saved and predictions generated.")
    except Exception as e:
        flash(f"Error: {str(e)}")
    return redirect(url_for('employee_dashboard'))

@app.route('/add_task', methods=['POST'])
def add_task():
    c.execute("INSERT INTO task VALUES (?,?,?,?)",
              (request.form['emp_id'], request.form['task'], "Pending", datetime.datetime.now().isoformat()))
    conn.commit()
    return redirect(url_for('employee_dashboard'))

@app.route('/complete_task', methods=['POST'])
def complete_task():
    c.execute("UPDATE task SET status='Done' WHERE emp_id=? AND task=?", 
              (request.form['emp_id'], request.form['task']))
    conn.commit()
    return redirect(url_for('employee_dashboard'))

@app.route('/chat', methods=['POST'])
def chat():
    emp_id = request.form['emp_id']
    user_msg = request.form['message']
    c.execute("INSERT INTO chat VALUES (?,?,?,?)", (emp_id, 'user', user_msg, datetime.datetime.now().isoformat()))
    c.execute("SELECT name, leaves_taken FROM employee WHERE id=?", (emp_id,))
    row = c.fetchone()
    name = row[0] if row else "Employee"
    leaves = row[1] if row else 0
    bot_reply = get_employee_bot_response(name, leaves, user_msg)
    c.execute("INSERT INTO chat VALUES (?,?,?,?)", (emp_id, 'bot', bot_reply, datetime.datetime.now().isoformat()))
    conn.commit()
    return redirect(url_for('employee_dashboard'))

@app.route('/download_pdf', methods=['POST'])
def download_pdf():
    try:
        emp_id = request.form['emp_id']
        c.execute("SELECT * FROM employee WHERE id=?", (emp_id,))
        profile = c.fetchone()
        if not profile:
            flash("Profile not found.")
            return redirect(url_for('employee_dashboard'))

        profile_dict = dict(zip(['id','name','age','income','sat','overtime','involve','feedback','leaves_taken','password_hash'], profile))
        c.execute("SELECT task, status FROM task WHERE emp_id=?", (emp_id,))
        tasks = c.fetchall()
        results = session.get('results', {})

        pdf_buffer = generate_pdf(emp_id, profile_dict, results, tasks)
        if not pdf_buffer or pdf_buffer.getvalue() == b'':
            flash("PDF generation failed. Try again.")
            return redirect(url_for('employee_dashboard'))

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"HR_Report_{emp_id}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        flash("PDF error. Check server logs.")
        print("CRITICAL PDF ERROR:", traceback.format_exc())
        return redirect(url_for('employee_dashboard'))

# -------------------------- Run --------------------------
if __name__ == '__main__':
    print("HR Insight Bot → http://127.0.0.1:5000")
    app.run(host='0.0.0.0', port=5000, debug=False)